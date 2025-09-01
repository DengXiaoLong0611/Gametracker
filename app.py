from fastapi import FastAPI, Request, HTTPException, Depends, status
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
import uvicorn
import os
import io
import csv
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path

logger = logging.getLogger(__name__)

from models import (
    Game, GameCreate, GameUpdate, LimitUpdate, GameStatus,
    Book, BookCreate, BookUpdate, BookResponse, ReadingCountResponse, BookStatus,
    User, UserCreate, UserLogin, UserResponse, Token,
    ExportRequest, ExportFormat
)
from store_adapter import GameStoreAdapter
from book_store import BookStore
from user_store import MultiUserStore
from database import db_manager, initialize_settings
from exceptions import GameTrackerException
from auth import (
    authenticate_user, create_access_token, get_current_active_user,
    get_password_hash, ACCESS_TOKEN_EXPIRE_MINUTES
)


from contextlib import asynccontextmanager

@asynccontextmanager
async def lifespan(app: FastAPI):
    # 启动时初始化
    if store.use_database:
        await db_manager.initialize()
        await db_manager.create_tables()
        async with db_manager.get_session() as session:
            await initialize_settings(session)
    yield
    # 关闭时清理
    if store.use_database:
        await db_manager.close()

# 全局store实例
store = GameStoreAdapter()
book_store = BookStore()
user_store = MultiUserStore()

# 同步创建app
def create_app_sync() -> FastAPI:
    """Create and configure the FastAPI application"""
    app = FastAPI(
        title="游戏追踪器",
        description="管理您的游戏进度",
        version="1.0.0",
        lifespan=lifespan
    )
    
    # 添加CORS中间件
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],  # 生产环境中应该设置具体的域名
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )
    
    # Configure static files and templates
    static_dir = Path("static")
    templates_dir = Path("templates")
    
    if static_dir.exists():
        app.mount("/static", StaticFiles(directory="static"), name="static")
    
    app.templates = Jinja2Templates(directory="templates") if templates_dir.exists() else None
    app.templates_dir_exists = templates_dir.exists()
    
    # 添加安全头
    @app.middleware("http")
    async def add_security_headers(request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        return response
    
    return app

app = create_app_sync()


@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    if app.templates_dir_exists:
        return app.templates.TemplateResponse("index.html", {"request": request})
    else:
        return HTMLResponse("""
        <html><body>
        <h1>Game Tracker API</h1>
        <p>API is running. Visit <a href="/docs">/docs</a> for API documentation.</p>
        </body></html>
        """)

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    """登录页面"""
    if app.templates_dir_exists:
        return app.templates.TemplateResponse("login.html", {"request": request})
    else:
        return HTMLResponse("""
        <html><body>
        <h1>Login Page</h1>
        <p>Please visit <a href="/docs">/docs</a> for API documentation.</p>
        </body></html>
        """)

@app.get("/reading", response_class=HTMLResponse)
async def reading_tracker(request: Request):
    if app.templates_dir_exists:
        return app.templates.TemplateResponse("reading.html", {"request": request})
    else:
        return HTMLResponse("""
        <html><body>
        <h1>Reading Tracker API</h1>
        <p>Reading tracker is running. Visit <a href="/docs">/docs</a> for API documentation.</p>
        </body></html>
        """)

@app.get("/health")
async def health_check():
    """健康检查端点"""
    return {
        "status": "healthy",
        "message": "游戏追踪器运行正常",
        "database_mode": store.use_database
    }

# ====================== 用户认证API ======================

@app.post("/api/auth/register", response_model=UserResponse)
async def register(user_data: UserCreate):
    """用户注册"""
    try:
        password_hash = get_password_hash(user_data.password)
        user = await user_store.create_user(user_data, password_hash)
        return UserResponse(
            id=user.id,
            username=user.username,
            email=user.email,
            is_active=user.is_active,
            created_at=user.created_at
        )
    except Exception as e:
        if "already exists" in str(e):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="邮箱已被注册"
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="注册失败"
        )

@app.post("/api/auth/login", response_model=Token)
async def login(user_credentials: UserLogin):
    """用户登录"""
    user = await authenticate_user(user_credentials.email, user_credentials.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="邮箱或密码错误",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token = create_access_token(
        data={"sub": str(user.id)},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    
    return Token(
        access_token=access_token,
        token_type="bearer",
        expires_in=ACCESS_TOKEN_EXPIRE_MINUTES * 60  # 转换为秒
    )

@app.get("/api/auth/me", response_model=UserResponse)
async def get_current_user_info(current_user: User = Depends(get_current_active_user)):
    """获取当前用户信息"""
    return UserResponse(
        id=current_user.id,
        username=current_user.username,
        email=current_user.email,
        is_active=current_user.is_active,
        created_at=current_user.created_at
    )

# ====================== 游戏管理API ======================

@app.get("/api/games")
async def get_games(current_user: User = Depends(get_current_active_user)):
    """获取当前用户的所有游戏"""
    try:
        if store.use_database:
            games = await user_store.get_all_games(current_user.id)
            return games
        else:
            # JSON模式暂时返回原有数据（待后续支持用户隔离）
            return await store.get_all_games()
    except GameTrackerException as e:
        logger.error(f"GameTrackerException in get_games: {str(e)}")
        raise e.to_http_exception() if hasattr(e, 'to_http_exception') else HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        error_str = str(e)
        logger.error(f"Unexpected error in get_games: {error_str}")
        
        # 检查是否是数据库模式问题（缺少user_id列）
        if "column" in error_str and "user_id" in error_str and "does not exist" in error_str:
            logger.warning("Database schema is outdated (missing user_id column), falling back to empty data")
            # 返回空的游戏数据结构，避免500错误
            return {
                "active": [],
                "paused": [],
                "casual": [],
                "planned": [],
                "finished": [],
                "dropped": []
            }
        
        raise HTTPException(status_code=500, detail=f"Internal server error: {error_str}")

@app.get("/api/active-count")
async def get_active_count(current_user: User = Depends(get_current_active_user)):
    """获取当前用户的活跃游戏计数"""
    try:
        if store.use_database:
            count_data = await user_store.get_active_count(current_user.id)
            return count_data
        else:
            return await store.get_active_count()
    except GameTrackerException as e:
        logger.error(f"GameTrackerException in get_active_count: {str(e)}")
        raise e.to_http_exception() if hasattr(e, 'to_http_exception') else HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        error_str = str(e)
        logger.error(f"Unexpected error in get_active_count: {error_str}")
        
        # 检查是否是数据库模式问题（缺少user_id列）
        if "column" in error_str and "user_id" in error_str and "does not exist" in error_str:
            logger.warning("Database schema is outdated (missing user_id column), falling back to empty count")
            # 返回空的计数数据结构，使用新用户的默认值
            return {
                "count": 0,
                "limit": 3,  # 新的默认限制
                "paused_count": 0,
                "casual_count": 0,
                "planned_count": 0
            }
        
        raise HTTPException(status_code=500, detail=f"Internal server error: {error_str}")

@app.post("/api/games", response_model=Game)
async def create_game(game: GameCreate, current_user: User = Depends(get_current_active_user)):
    """添加新游戏"""
    try:
        if store.use_database:
            return await user_store.add_game(current_user.id, game)
        else:
            return await store.add_game(game)
    except GameTrackerException as e:
        logger.error(f"GameTrackerException in create_game: {str(e)}")
        raise e.to_http_exception() if hasattr(e, 'to_http_exception') else HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        error_str = str(e)
        logger.error(f"Unexpected error in create_game: {error_str}")
        
        # 检查是否是数据库模式问题
        if "column" in error_str and "user_id" in error_str and "does not exist" in error_str:
            logger.warning("Database schema is outdated (missing user_id column), cannot create game")
            raise HTTPException(status_code=503, detail="Database schema needs to be updated. Please contact administrator.")
        
        raise HTTPException(status_code=500, detail=f"Internal server error: {error_str}")

@app.patch("/api/games/{game_id}", response_model=Game)
async def update_game(game_id: int, updates: GameUpdate, current_user: User = Depends(get_current_active_user)):
    """更新游戏"""
    try:
        if store.use_database:
            return await user_store.update_game(current_user.id, game_id, updates)
        else:
            return await store.update_game(game_id, updates)
    except GameTrackerException as e:
        raise e.to_http_exception() if hasattr(e, 'to_http_exception') else HTTPException(status_code=500, detail=str(e))

@app.delete("/api/games/{game_id}")
async def delete_game(game_id: int, current_user: User = Depends(get_current_active_user)):
    """删除游戏"""
    try:
        if store.use_database:
            success = await user_store.delete_game(current_user.id, game_id)
        else:
            success = await store.delete_game(game_id)
        return {"success": success}
    except GameTrackerException as e:
        raise e.to_http_exception() if hasattr(e, 'to_http_exception') else HTTPException(status_code=500, detail=str(e))

@app.post("/api/settings/limit")
async def update_limit(limit_data: LimitUpdate, current_user: User = Depends(get_current_active_user)):
    """更新游戏数量限制"""
    try:
        if store.use_database:
            await user_store.update_game_limit(current_user.id, limit_data.limit)
        else:
            await store.update_limit(limit_data.limit)
        return {"message": "游戏限制更新成功", "new_limit": limit_data.limit}
    except GameTrackerException as e:
        raise e.to_http_exception() if hasattr(e, 'to_http_exception') else HTTPException(status_code=500, detail=str(e))

# ================== 书籍阅读追踪器 API ==================

@app.get("/api/books", response_model=BookResponse)
async def get_books(current_user: User = Depends(get_current_active_user)):
    """获取当前用户的所有书籍，按状态分组"""
    try:
        if store.use_database:
            books_data = await user_store.get_all_books(current_user.id)
            return BookResponse(**books_data)
        else:
            books_data = book_store.get_all_books()
            return BookResponse(**books_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reading-count", response_model=ReadingCountResponse)
async def get_reading_count(current_user: User = Depends(get_current_active_user)):
    """获取当前用户的阅读数量和限制"""
    try:
        if store.use_database:
            count_data = await user_store.get_reading_count(current_user.id)
        else:
            count_data = book_store.get_reading_count()
        return ReadingCountResponse(**count_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/books", response_model=Book)
async def create_book(book: BookCreate, current_user: User = Depends(get_current_active_user)):
    """创建新书籍"""
    try:
        if store.use_database:
            return await user_store.add_book(current_user.id, book)
        else:
            return book_store.add_book(book)
    except Exception as e:
        error_str = str(e)
        logger.error(f"Error in create_book: {error_str}")
        
        # 检查是否是数据库模式问题
        if "column" in error_str and "user_id" in error_str and "does not exist" in error_str:
            logger.warning("Database schema is outdated for books (missing user_id column)")
            raise HTTPException(status_code=503, detail="Database schema needs to be updated. Please contact administrator.")
        
        raise HTTPException(status_code=400, detail=str(e))

@app.patch("/api/books/{book_id}", response_model=Book)
async def update_book(book_id: int, updates: BookUpdate, current_user: User = Depends(get_current_active_user)):
    """更新书籍信息"""
    try:
        if store.use_database:
            return await user_store.update_book(current_user.id, book_id, updates)
        else:
            return book_store.update_book(book_id, updates)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/api/books/{book_id}")
async def delete_book(book_id: int, current_user: User = Depends(get_current_active_user)):
    """删除书籍"""
    try:
        if store.use_database:
            success = await user_store.delete_book(current_user.id, book_id)
        else:
            success = book_store.delete_book(book_id)
        return {"success": success}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/books/settings/limit")
async def update_reading_limit(limit_data: LimitUpdate, current_user: User = Depends(get_current_active_user)):
    """更新阅读数量限制"""
    try:
        if store.use_database:
            await user_store.update_book_limit(current_user.id, limit_data.limit)
        else:
            book_store.update_limit(limit_data.limit)
        return {"message": "阅读限制更新成功", "new_limit": limit_data.limit}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ====================== 数据导出API ======================

@app.post("/api/export")
async def export_user_data(
    export_request: ExportRequest,
    current_user: User = Depends(get_current_active_user)
):
    """导出用户数据"""
    try:
        # 收集用户数据
        export_data = {
            "user_info": {
                "username": current_user.username,
                "email": current_user.email,
                "export_date": datetime.now().isoformat()
            }
        }
        
        # 导出游戏数据
        if export_request.include_games:
            if store.use_database:
                games_data = await user_store.get_all_games(current_user.id)
            else:
                games_data = await store.get_all_games()
            export_data["games"] = games_data
        
        # 导出书籍数据
        if export_request.include_books:
            if store.use_database:
                books_data = await user_store.get_all_books(current_user.id)
            else:
                books_data = book_store.get_all_books()
            export_data["books"] = books_data
        
        # 根据格式返回数据
        if export_request.format == ExportFormat.JSON:
            return JSONResponse(
                content=export_data,
                headers={
                    "Content-Disposition": f"attachment; filename=game_tracker_export_{current_user.username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                }
            )
        
        elif export_request.format == ExportFormat.CSV:
            return await _generate_csv_export(export_data, current_user.username)
        
        elif export_request.format == ExportFormat.EXCEL:
            return await _generate_excel_export(export_data, current_user.username)
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")

async def _generate_csv_export(export_data: dict, username: str):
    """生成CSV格式导出"""
    import io
    import csv
    
    output = io.StringIO()
    
    # 导出游戏数据
    if "games" in export_data:
        output.write("=== 游戏数据 ===\n")
        writer = csv.writer(output)
        writer.writerow(["游戏名称", "状态", "评分", "笔记", "理由", "创建时间", "结束时间"])
        
        for status, games in export_data["games"].items():
            for game in games:
                writer.writerow([
                    game.get("name", ""),
                    game.get("status", ""),
                    game.get("rating", ""),
                    game.get("notes", ""),
                    game.get("reason", ""),
                    game.get("created_at", ""),
                    game.get("ended_at", "")
                ])
        output.write("\n")
    
    # 导出书籍数据
    if "books" in export_data:
        output.write("=== 书籍数据 ===\n")
        writer = csv.writer(output)
        writer.writerow(["书名", "作者", "状态", "进度", "评分", "笔记", "理由", "创建时间", "结束时间"])
        
        for status, books in export_data["books"].items():
            for book in books:
                writer.writerow([
                    book.get("title", ""),
                    book.get("author", ""),
                    book.get("status", ""),
                    book.get("progress", ""),
                    book.get("rating", ""),
                    book.get("notes", ""),
                    book.get("reason", ""),
                    book.get("created_at", ""),
                    book.get("ended_at", "")
                ])
    
    csv_content = output.getvalue()
    output.close()
    
    from fastapi.responses import Response
    return Response(
        content=csv_content.encode('utf-8-sig'),  # 使用BOM确保Excel正确显示中文
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename=game_tracker_export_{username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

async def _generate_excel_export(export_data: dict, username: str):
    """生成Excel格式导出"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import io
        
        wb = Workbook()
        
        # 删除默认工作表
        wb.remove(wb.active)
        
        # 创建游戏数据工作表
        if "games" in export_data:
            ws_games = wb.create_sheet("游戏数据")
            
            # 标题行
            headers = ["游戏名称", "状态", "评分", "笔记", "理由", "创建时间", "结束时间"]
            for col, header in enumerate(headers, 1):
                cell = ws_games.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 数据行
            row = 2
            for status, games in export_data["games"].items():
                for game in games:
                    ws_games.cell(row=row, column=1, value=game.get("name", ""))
                    ws_games.cell(row=row, column=2, value=game.get("status", ""))
                    ws_games.cell(row=row, column=3, value=game.get("rating", ""))
                    ws_games.cell(row=row, column=4, value=game.get("notes", ""))
                    ws_games.cell(row=row, column=5, value=game.get("reason", ""))
                    ws_games.cell(row=row, column=6, value=game.get("created_at", ""))
                    ws_games.cell(row=row, column=7, value=game.get("ended_at", ""))
                    row += 1
        
        # 创建书籍数据工作表
        if "books" in export_data:
            ws_books = wb.create_sheet("书籍数据")
            
            # 标题行
            headers = ["书名", "作者", "状态", "进度", "评分", "笔记", "理由", "创建时间", "结束时间"]
            for col, header in enumerate(headers, 1):
                cell = ws_books.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 数据行
            row = 2
            for status, books in export_data["books"].items():
                for book in books:
                    ws_books.cell(row=row, column=1, value=book.get("title", ""))
                    ws_books.cell(row=row, column=2, value=book.get("author", ""))
                    ws_books.cell(row=row, column=3, value=book.get("status", ""))
                    ws_books.cell(row=row, column=4, value=book.get("progress", ""))
                    ws_books.cell(row=row, column=5, value=book.get("rating", ""))
                    ws_books.cell(row=row, column=6, value=book.get("notes", ""))
                    ws_books.cell(row=row, column=7, value=book.get("reason", ""))
                    ws_books.cell(row=row, column=8, value=book.get("created_at", ""))
                    ws_books.cell(row=row, column=9, value=book.get("ended_at", ""))
                    row += 1
        
        # 保存到内存
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        from fastapi.responses import Response
        return Response(
            content=excel_file.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=game_tracker_export_{username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel导出功能需要安装openpyxl库")

# ====================== 数据迁移API ======================

@app.post("/api/admin/migrate-schema")
async def migrate_schema_only():
    """单独运行数据库模式迁移（调试用）"""
    try:
        success = await _migrate_database_schema_direct()
        if success:
            return {"success": True, "message": "数据库模式迁移完成"}
        else:
            return {"success": False, "message": "数据库模式迁移失败"}
    except Exception as e:
        return {"success": False, "message": f"迁移异常: {str(e)}"}

@app.post("/api/admin/force-migrate")
async def force_migrate_schema():
    """强制执行数据库模式迁移（简化版本）"""
    try:
        from sqlalchemy import text
        
        async with db_manager.get_session() as session:
            migration_log = []
            migration_log.append("开始强制数据库模式迁移...")
            
            # 1. 首先确保有用户表
            try:
                await session.execute(text("""
                    CREATE TABLE IF NOT EXISTS users (
                        id SERIAL PRIMARY KEY,
                        username VARCHAR(50) NOT NULL,
                        email VARCHAR(100) UNIQUE NOT NULL,
                        password_hash VARCHAR(255) NOT NULL,
                        is_active BOOLEAN DEFAULT true NOT NULL,
                        created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP NOT NULL
                    );
                """))
                migration_log.append("✅ 用户表检查/创建完成")
                await session.commit()
            except Exception as e:
                migration_log.append(f"❌ 用户表操作失败: {str(e)}")
                await session.rollback()
            
            # 2. 创建默认用户
            try:
                result = await session.execute(text("""
                    INSERT INTO users (username, email, password_hash) 
                    VALUES ('default_user', 'default@gametracker.com', '$2b$12$defaulthash') 
                    ON CONFLICT (email) DO NOTHING
                    RETURNING id;
                """))
                user_id = result.scalar()
                if user_id:
                    migration_log.append(f"✅ 创建默认用户 ID: {user_id}")
                else:
                    # 获取现有用户ID
                    existing = await session.execute(text("SELECT id FROM users WHERE email = 'default@gametracker.com' LIMIT 1"))
                    user_id = existing.scalar() or 1
                    migration_log.append(f"✅ 使用现有默认用户 ID: {user_id}")
                await session.commit()
            except Exception as e:
                migration_log.append(f"❌ 默认用户操作失败: {str(e)}")
                await session.rollback()
                user_id = 1  # 后备用户ID
            
            # 3. 为games表添加user_id列（如果不存在）
            try:
                await session.execute(text(f"""
                    DO $$
                    BEGIN
                        IF NOT EXISTS (SELECT 1 FROM information_schema.columns 
                                      WHERE table_name='games' AND column_name='user_id') THEN
                            ALTER TABLE games ADD COLUMN user_id INTEGER NOT NULL DEFAULT {user_id};
                            ALTER TABLE games ADD CONSTRAINT fk_games_user_id 
                                FOREIGN KEY (user_id) REFERENCES users(id);
                            CREATE INDEX ix_games_user_id ON games (user_id);
                        END IF;
                    END $$;
                """))
                migration_log.append("✅ games表user_id列操作完成")
                await session.commit()
            except Exception as e:
                migration_log.append(f"❌ games表操作失败: {str(e)}")
                await session.rollback()
            
            # 4. 创建settings表
            try:
                await session.execute(text("""
                    CREATE TABLE IF NOT EXISTS settings (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL REFERENCES users(id),
                        key VARCHAR(50) NOT NULL,
                        value INTEGER NOT NULL,
                        created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP NOT NULL,
                        UNIQUE(user_id, key)
                    );
                """))
                migration_log.append("✅ settings表操作完成")
                await session.commit()
            except Exception as e:
                migration_log.append(f"❌ settings表操作失败: {str(e)}")
                await session.rollback()
            
            migration_log.append("🎉 强制迁移完成")
            return {
                "success": True, 
                "message": "强制数据库模式迁移完成",
                "log": migration_log
            }
            
    except Exception as e:
        return {
            "success": False, 
            "message": f"强制迁移失败: {str(e)}",
            "log": migration_log if 'migration_log' in locals() else []
        }

async def _migrate_database_schema_direct():
    """直接进行数据库模式迁移，不依赖migrate_database_schema模块"""
    try:
        from sqlalchemy import text
        
        async with db_manager.get_session() as session:
            logger.info("开始直接数据库模式迁移...")
            
            # 检查用户表是否存在
            try:
                users_table_check = await session.execute(text("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_name = 'users'
                    );
                """))
                users_table_exists = users_table_check.scalar()
                logger.info(f"用户表存在检查: {users_table_exists}")
            except Exception as e:
                logger.error(f"检查用户表失败: {str(e)}")
                users_table_exists = False
            
            if not users_table_exists:
                try:
                    logger.info("创建用户表...")
                    await session.execute(text("""
                        CREATE TABLE users (
                            id SERIAL PRIMARY KEY,
                            username VARCHAR(50) NOT NULL,
                            email VARCHAR(100) UNIQUE NOT NULL,
                            password_hash VARCHAR(255) NOT NULL,
                            is_active BOOLEAN DEFAULT true NOT NULL,
                            created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP NOT NULL,
                            CONSTRAINT username_min_length CHECK (LENGTH(TRIM(username)) >= 2),
                            CONSTRAINT email_not_empty CHECK (LENGTH(TRIM(email)) > 0)
                        );
                    """))
                    
                    # 创建索引
                    await session.execute(text("CREATE INDEX ix_users_id ON users (id);"))
                    await session.execute(text("CREATE INDEX ix_users_email ON users (email);"))
                    logger.info("✅ 用户表创建成功")
                    await session.commit()  # 立即提交用户表创建
                except Exception as e:
                    logger.error(f"创建用户表失败: {str(e)}")
                    await session.rollback()
                    return False
            else:
                logger.info("✅ 用户表已存在")
            
            # 检查games表的user_id列是否存在
            try:
                games_user_id_check = await session.execute(text("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.columns 
                        WHERE table_name = 'games' AND column_name = 'user_id'
                    );
                """))
                games_user_id_exists = games_user_id_check.scalar()
                logger.info(f"games表user_id列存在检查: {games_user_id_exists}")
            except Exception as e:
                logger.error(f"检查games表user_id列失败: {str(e)}")
                games_user_id_exists = False
            
            if not games_user_id_exists:
                try:
                    logger.info("为games表添加user_id列...")
                    
                    # 创建默认用户（如果需要）
                    default_user_check = await session.execute(text("""
                        SELECT id FROM users WHERE email = 'default@gametracker.com' LIMIT 1;
                    """))
                    default_user_id = default_user_check.scalar()
                    
                    if not default_user_id:
                        logger.info("创建默认用户...")
                        result = await session.execute(text("""
                            INSERT INTO users (username, email, password_hash) 
                            VALUES ('default_user', 'default@gametracker.com', '$2b$12$defaulthash') 
                            RETURNING id;
                        """))
                        default_user_id = result.scalar()
                        logger.info(f"✅ 默认用户创建成功，ID: {default_user_id}")
                        await session.commit()  # 立即提交默认用户创建
                    
                    # 添加user_id列
                    await session.execute(text(f"""
                        ALTER TABLE games ADD COLUMN user_id INTEGER NOT NULL DEFAULT {default_user_id};
                    """))
                    logger.info("games表user_id列添加完成")
                    
                    # 添加外键约束
                    await session.execute(text("""
                        ALTER TABLE games ADD CONSTRAINT fk_games_user_id 
                        FOREIGN KEY (user_id) REFERENCES users(id);
                    """))
                    logger.info("games表外键约束添加完成")
                    
                    # 创建索引
                    await session.execute(text("CREATE INDEX ix_games_user_id ON games (user_id);"))
                    logger.info("✅ games表user_id列、外键和索引添加成功")
                    await session.commit()  # 立即提交games表修改
                except Exception as e:
                    logger.error(f"为games表添加user_id列失败: {str(e)}")
                    await session.rollback()
                    return False
            else:
                logger.info("✅ games表已有user_id列")
            
            # 检查settings表
            try:
                settings_table_check = await session.execute(text("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_name = 'settings'
                    );
                """))
                settings_table_exists = settings_table_check.scalar()
                logger.info(f"settings表存在检查: {settings_table_exists}")
            except Exception as e:
                logger.error(f"检查settings表失败: {str(e)}")
                settings_table_exists = False
            
            if not settings_table_exists:
                try:
                    logger.info("创建settings表...")
                    await session.execute(text("""
                        CREATE TABLE settings (
                            id SERIAL PRIMARY KEY,
                            user_id INTEGER NOT NULL REFERENCES users(id),
                            key VARCHAR(50) NOT NULL,
                            value INTEGER NOT NULL,
                            created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP NOT NULL,
                            UNIQUE(user_id, key)
                        );
                    """))
                    
                    await session.execute(text("CREATE INDEX ix_settings_id ON settings (id);"))
                    await session.execute(text("CREATE INDEX ix_settings_user_id ON settings (user_id);"))
                    logger.info("✅ settings表创建成功")
                    await session.commit()  # 立即提交settings表创建
                except Exception as e:
                    logger.error(f"创建settings表失败: {str(e)}")
                    await session.rollback()
                    return False
            else:
                logger.info("✅ settings表已存在")
            
            logger.info("🎉 直接数据库模式迁移完成!")
            return True
            
    except Exception as e:
        logger.error(f"❌ 直接数据库迁移失败: {str(e)}")
        return False

@app.post("/api/admin/migrate-legacy-data")
async def migrate_legacy_data(current_user: User = Depends(get_current_active_user)):
    """迁移遗留数据到当前用户账户 (仅限hero19950611用户)"""
    
    # 安全检查：仅允许特定用户执行迁移
    if current_user.email != "382592406@qq.com" or current_user.username != "hero19950611":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只有原账户持有者可以执行数据迁移"
        )
    
    try:
        from pathlib import Path
        import json
        
        migration_result = {
            "games_migrated": 0,
            "books_migrated": 0,
            "errors": []
        }
        
        # 首先运行数据库模式迁移（处理缺少user_id列的情况）
        try:
            migration_success = await _migrate_database_schema_direct()
            if not migration_success:
                logger.warning("数据库模式迁移失败，但继续尝试数据迁移")
            else:
                logger.info("数据库模式迁移完成")
        except Exception as schema_error:
            logger.error(f"数据库模式迁移异常: {str(schema_error)}")
            # 继续执行，可能数据库已经是最新模式了
        
        # 检查是否已经迁移过
        try:
            existing_games = await user_store.get_all_games(current_user.id)
            total_existing = sum(len(games) for games in existing_games.values())
            
            if total_existing > 0:
                return {
                    "success": False,
                    "message": "数据已存在，避免重复迁移",
                    "existing_games": total_existing
                }
        except Exception as e:
            # 如果仍然无法访问游戏数据，说明有其他问题
            logger.error(f"无法检查现有游戏数据: {str(e)}")
            return {
                "success": False,
                "message": f"数据库访问失败: {str(e)}"
            }
        
        # 迁移游戏数据
        games_file = Path("games_data.json")
        if games_file.exists():
            with open(games_file, 'r', encoding='utf-8') as f:
                games_data = json.load(f)
            
            # 处理旧格式数据
            if isinstance(games_data.get('games'), dict):
                for game_id, game_data in games_data['games'].items():
                    try:
                        # 转换状态
                        status_str = game_data.get('status', 'active')
                        game_status = getattr(GameStatus, status_str.upper(), GameStatus.ACTIVE)
                        
                        game_create = GameCreate(
                            name=game_data.get('name', ''),
                            status=game_status,
                            notes=game_data.get('notes', ''),
                            rating=game_data.get('rating'),
                            reason=game_data.get('reason', '')
                        )
                        
                        await user_store.add_game(current_user.id, game_create)
                        migration_result["games_migrated"] += 1
                        
                    except Exception as e:
                        migration_result["errors"].append(f"游戏迁移失败: {game_data.get('name', 'Unknown')} - {str(e)}")
        
        # 迁移书籍数据
        books_file = Path("books_data.json")
        if books_file.exists():
            with open(books_file, 'r', encoding='utf-8') as f:
                books_data = json.load(f)
            
            # 处理书籍数据
            if isinstance(books_data.get('books'), dict):
                for book_id, book_data in books_data['books'].items():
                    try:
                        # 转换状态
                        status_str = book_data.get('status', 'reading')
                        book_status = getattr(BookStatus, status_str.upper(), BookStatus.READING)
                        
                        book_create = BookCreate(
                            title=book_data.get('title', ''),
                            author=book_data.get('author', ''),
                            status=book_status,
                            notes=book_data.get('notes', ''),
                            rating=book_data.get('rating'),
                            reason=book_data.get('reason', ''),
                            progress=book_data.get('progress', '')
                        )
                        
                        await user_store.add_book(current_user.id, book_create)
                        migration_result["books_migrated"] += 1
                        
                    except Exception as e:
                        migration_result["errors"].append(f"书籍迁移失败: {book_data.get('title', 'Unknown')} - {str(e)}")
        
        return {
            "success": True,
            "message": "数据迁移完成！",
            "games_migrated": migration_result["games_migrated"],
            "books_migrated": migration_result["books_migrated"],
            "errors": migration_result["errors"]
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"迁移过程中发生错误: {str(e)}"
        )

# GitHub同步相关API端点
@app.get("/api/sync/status")
async def get_sync_status():
    """获取GitHub同步状态"""
    try:
        # 调用同步版本的方法
        if store.use_database:
            # 数据库模式暂不支持GitHub同步
            return {
                "enabled": False,
                "reason": "GitHub同步仅在JSON存储模式下可用"
            }
        else:
            return store._store.get_sync_status()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取同步状态失败: {str(e)}")

@app.post("/api/sync/to-github")
async def sync_to_github():
    """手动同步数据到GitHub"""
    try:
        if store.use_database:
            raise HTTPException(status_code=400, detail="数据库模式下不支持GitHub同步")
        
        success = store._store.manual_sync_to_github()
        if success:
            return {"success": True, "message": "数据已成功同步到GitHub"}
        else:
            return {"success": False, "message": "同步到GitHub失败，请检查配置"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"同步失败: {str(e)}")

@app.post("/api/sync/from-github")
async def sync_from_github():
    """手动从GitHub同步数据"""
    try:
        if store.use_database:
            raise HTTPException(status_code=400, detail="数据库模式下不支持GitHub同步")
        
        success = store._store.manual_sync_from_github()
        if success:
            return {"success": True, "message": "数据已成功从GitHub同步"}
        else:
            return {"success": False, "message": "从GitHub同步失败，请检查配置"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"同步失败: {str(e)}")

if __name__ == "__main__":
    # 环境检测和配置
    deployment_env = os.getenv("DEPLOYMENT_ENV", "local")
    host = os.getenv("HOST", "0.0.0.0")
    debug = os.getenv("DEBUG", "False").lower() == "true"
    
    # 端口配置根据部署环境
    if deployment_env == "tencent-container":
        # 腾讯云容器环境
        port = int(os.getenv("PORT", "9000"))
    elif deployment_env == "tencent-scf":
        # 腾讯云云函数环境
        port = int(os.getenv("PORT", "9000"))
    else:
        # 本地开发环境
        port = int(os.getenv("PORT", "8001"))
    
    print(f"Starting Game & Reading Tracker")
    print(f"Environment: {deployment_env}")
    print(f"Server: {host}:{port}")
    print(f"Static files: {Path('static').exists()}")
    print(f"Templates: {Path('templates').exists()}")
    print(f"Debug mode: {debug}")
    
    uvicorn.run(
        "app:app",
        host=host,
        port=port,
        reload=debug,
        access_log=True,
        log_level="info" if not debug else "debug"
    )